"""
Excel — create formatted spreadsheets with openpyxl (headers, styling,
formulas), open them visibly in Excel via win32com, and read/summarize
existing workbooks on command.
"""
import re
import time
from difflib import get_close_matches
from pathlib import Path

from config import Config
from brain.prompts import EXCEL_CREATE_PROMPT


def _slug(text):
    s = re.sub(r"[^\w\s-]", "", text)[:60].strip().replace(" ", "_")
    return s or "spreadsheet"


def _open_in_excel(path, ctx):
    try:
        import win32com.client as win32
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = True
        excel.Workbooks.Open(str(path))

        def _closer(x=excel):
            try:
                x.Quit()
            except Exception:
                pass
        ctx.registry.register("document", Path(path).name,
                              window_title=Path(path).stem, closer=_closer,
                              extra={"path": str(path)})
        return True
    except Exception:
        try:
            import os
            os.startfile(str(path))
            ctx.registry.register("document", Path(path).name,
                                  window_title=Path(path).stem,
                                  extra={"path": str(path)})
        except Exception:
            pass
        return False


def _find_workbook(name):
    name_l = (name or "").lower().replace(".xlsx", "").replace(".xls", "")
    candidates = []
    try:
        for p in Config.DESKTOP_PATH.glob("*.xls*"):
            if p.suffix.lower() in (".xlsx", ".xls", ".xlsm", ".csv"):
                candidates.append(p)
    except Exception:
        pass
    if not candidates:
        return None
    for p in candidates:
        if name_l in p.stem.lower():
            return p
    matches = get_close_matches(name_l, [p.stem.lower() for p in candidates],
                                n=1, cutoff=0.4)
    if matches:
        for p in candidates:
            if p.stem.lower() == matches[0]:
                return p
    return None


# ---------------------------------------------------------------------------
# Create
# ---------------------------------------------------------------------------
def create_spreadsheet(topic, ctx):
    topic = (topic or "").strip()
    if not topic:
        return "A spreadsheet of what, sir?"
    if not ctx.llm.available:
        return "My design brain needs the OpenRouter key, sir."

    ctx.speaker.speak(f"Designing the spreadsheet, sir.")
    data = ctx.llm.quick_json(EXCEL_CREATE_PROMPT.format(topic=topic),
                              max_tokens=3000)
    if not data or not data.get("sheets"):
        return "I couldn't design that spreadsheet, sir. Try phrasing it differently."

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for spec in data["sheets"][:6]:
            name = re.sub(r"[\\/*?\[\]:]", "", str(spec.get("name", "Sheet")))[:31] or "Sheet"
            ws = wb.create_sheet(title=name)
            headers = spec.get("headers") or []
            rows = spec.get("rows") or []
            formulas = spec.get("formulas") or []

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=str(h))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            for r, row in enumerate(rows, 2):
                for c, val in enumerate(row, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = border

            for f in formulas:
                try:
                    cell_ref = str(f.get("cell", "")).strip()
                    formula = str(f.get("formula", "")).strip()
                    if cell_ref and formula.startswith("="):
                        ws[cell_ref] = formula
                except Exception:
                    continue

            for c in range(1, len(headers) + 1):
                longest = len(str(headers[c - 1])) if c <= len(headers) else 8
                for r in range(2, min(len(rows) + 2, 30)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        longest = max(longest, len(str(v)))
                ws.column_dimensions[get_column_letter(c)].width = min(longest + 3, 42)

            ws.freeze_panes = "A2"

        title = re.sub(r"[\\/*?\[\]:]", "", str(data.get("title", topic)))[:40]
        path = Config.DESKTOP_PATH / f"{_slug(title)}.xlsx"
        wb.save(str(path))
    except Exception as exc:
        return f"Spreadsheet construction failed: {exc}."

    _open_in_excel(path, ctx)
    n_sheets = len(data["sheets"])
    return (f"Done, sir. {path.name} is on your desktop and open in Excel — "
            f"{n_sheets} sheet{'s' if n_sheets != 1 else ''}.")


# ---------------------------------------------------------------------------
# Read / summarize
# ---------------------------------------------------------------------------
def read_workbook(name, ctx):
    path = _find_workbook(name)
    if path is None:
        return f"I can't find a spreadsheet called {name} on your desktop, sir."
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        parts = []
        for ws in wb.worksheets[:4]:
            dims = f"{ws.max_row} rows by {ws.max_column} columns"
            preview = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 5:
                    break
                preview.append(", ".join("" if v is None else str(v) for v in row[:8]))
            parts.append(f"Sheet '{ws.title}' — {dims}. First rows: " +
                         " | ".join(preview))
        wb.close()
    except Exception as exc:
        return f"I couldn't read {path.name}: {exc}."

    raw = "\n".join(parts)
    summary = ctx.llm.quick(
        f"Summarize this spreadsheet content in 3 spoken sentences — what it "
        f"contains and anything notable:\n\n{raw}",
        max_tokens=220) if ctx.llm.available else ""
    if summary:
        return summary
    return (f"{path.name}: " + " ".join(parts))[:600]


# ---------------------------------------------------------------------------
# Skill dispatch entry
# ---------------------------------------------------------------------------
def handle(intent, ctx):
    skill = intent.get("skill")
    params = intent.get("params", {}) or {}
    if skill == "excel.create":
        return create_spreadsheet(params.get("topic", ""), ctx)
    if skill == "excel.read":
        return read_workbook(params.get("file", ""), ctx)
    return None
